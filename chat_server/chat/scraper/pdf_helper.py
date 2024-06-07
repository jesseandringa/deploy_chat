import os
import re

import fitz
import openpyxl


def chunk_xlsx_into_paragraphs(xlsx_path):
    """Chunks an XLSX file into paragraphs."""
    chunks = []

    # Open the XLSX file
    xlsx_document = openpyxl.load_workbook(xlsx_path)
    sheet = xlsx_document.active

    paragraph_number = 0
    labels = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, max_col=sheet.max_column
    ):
        paragraph = ""
        for i, cell in enumerate(row):
            if paragraph_number == 0:
                labels.append(cell.value)
            elif cell.value:
                paragraph += str(labels[i]) + ": " + str(cell.value) + ", "
        if paragraph_number > 0:
            chunk = [xlsx_path, paragraph_number, paragraph]
            chunks.append(chunk)
        paragraph_number += 1
    return chunks


def chunk_pdf_into_paragraphs(pdf_path):
    """Chunks a PDF into paragraphs."""
    chunks = []

    # Open the PDF file
    pdf_document = fitz.open(pdf_path)

    paragraph_number = 0

    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        text = page.get_text("text")
        # text_length = len(text)
        paragraphs = re.split(r"(?<=[.!?])\s*\n", text)
        # if len(paragraphs) < text_length / 500:
        #     extra_splitted = []
        #     for paragraph in paragraphs:
        #         # split paragraph by single newline but keep paragraphs one list
        #         extra_splitted.extend(paragraph.split("\n"))
        #     paragraphs = extra_splitted

        for paragraph in paragraphs:
            if paragraph.strip():  # Only process non-empty paragraphs
                paragraph_number += 1
                chunk = [pdf_path, page_num, paragraph_number, paragraph.strip()]
                chunks.append(chunk)
                # print("paragraph:", paragraph.strip())
        # print("paragraph_number:", paragraph_number)
    return chunks


def get_all_pdfs(folder_path):
    """Get all PDF files in a folder."""
    pdf_files = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".pdf"):
                pdf_files.append(os.path.join(root, file))
    return pdf_files


def get_all_xlsx(folder_path):
    """Get all XLSX files in a folder."""
    xlsx_files = []
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx"):
                xlsx_files.append(os.path.join(root, file))
    return xlsx_files


def insert_all_file_chunks_to_database(folder_path, db):
    """Insert all PDF chunks to the database."""
    print("inserting all file chunks to database")
    pdf_files = get_all_pdfs(folder_path)
    xlsx_files = get_all_xlsx(folder_path)
    print("pdf_files:", len(pdf_files))
    for pdf_file in pdf_files:
        print("inserting pdf:")
        chunks = chunk_pdf_into_paragraphs(pdf_file)
        for i, chunk in enumerate(chunks):
            text = chunk[2]
            if len(chunk[2]) < 50:
                text = chunks[i - 1][2] + " " + chunk[2] + " " + chunks[i + 1][2]
            db.insert_data_into_pdf_text_table(str(chunk[0]), int(chunk[1]), str(text))

    for xlsx_file in xlsx_files:
        chunks = chunk_xlsx_into_paragraphs(xlsx_file)
        for i, chunk in enumerate(chunks):
            db.insert_data_into_pdf_text_table(
                str(chunk[0]), int(chunk[1]), str(chunk[2])
            )


# pdf_path = "chat_server/chat/file_resources/murray-muni-resources/https:$$$$$$codelibrary.amlegal.com$$$codes$$$murrayut$$$latest$$$murray_ut$$$0-0-0-14#codecontent.pdf"
# pdf_path = "chat_server/chat/file_resources/gunnison-resources/https:$$$$$$www.gunnisoncounty.org$$$AgendaCenter$$$ViewFile$$$Agenda$$$_01022024-953.pdf"
# chunk_pdf_into_paragraphs(pdf_path, "end_sentence_newline")
# xl_path = "chat_server/chat/file_resources/gunnison-resources/https:$$$$$$www.gunnisoncounty.org$$$DocumentCenter$$$View$$$14027$$$Public-Data---VALUES-2824.xlsx"
# chunks = chunk_xlsx_into_paragraphs(xl_path)
# for i in range(10):
#     print(chunks[i])
